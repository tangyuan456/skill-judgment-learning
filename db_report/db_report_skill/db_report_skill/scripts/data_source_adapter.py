#!/usr/bin/env python3
"""
data_source_adapter.py — 统一数据接入

根据 intent.data_source_type 调用不同适配器，输出标准 records.json。
适配器：
  - local_file: 本地 .log/.xlsx/.json/.csv
  - local_data: 粘贴的 JSON 文本

标准结构见 references/数据源接入规范.md §1。
"""
import argparse
import csv
import json
import os
import re
import sys
from collections import defaultdict
from typing import Any, Dict, List, Optional

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

# 懒加载上游 log 解析（仅在处理 .log 时才需要；寻找 tdsql-b-whitepaper/scripts）
_UPSTREAM_CANDIDATES = [
    os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))),
                 'tdsql-b-whitepaper', 'scripts'),
    os.path.join(os.getcwd(), 'tdsql-b-whitepaper', 'scripts'),
    os.path.join(os.getcwd(), '..', 'tdsql-b-whitepaper', 'scripts'),
]


def _load_parse_log():
    """动态加载 tdsql-b-whitepaper/scripts/log_to_excel.parse_log。"""
    for cand in _UPSTREAM_CANDIDATES:
        cand = os.path.abspath(cand)
        if os.path.exists(os.path.join(cand, 'log_to_excel.py')):
            if cand not in sys.path:
                sys.path.insert(0, cand)
            from log_to_excel import parse_log  # type: ignore
            return parse_log
    raise RuntimeError(
        '未找到 tdsql-b-whitepaper/scripts/log_to_excel.py（解析 .log 文件需要此脚本）'
    )


# ============== 反向场景映射 ==============
SCEN_FROM_DB = {
    'point_select': 'oltp_point_select',
    'read_only': 'oltp_read_only',
    'write_only': 'oltp_write_only',
    'read_write': 'oltp_read_write',
    'update_index': 'oltp_update_index',
    'update_non_index': 'oltp_update_non_index',
    'random_points': 'oltp_random_points',
    'random_ranges': 'oltp_random_ranges',
    'tpmC': 'tpmC',
    'NewOrders': 'tpmC',
}


def _normalize_scenario(name: str) -> Optional[str]:
    """从 test_name 字符串中识别 sysbench 标准场景名。"""
    if not name:
        return None
    # 已知标准场景枚举（按最长优先匹配，避免 read_write 被 read_only 误匹配）
    KNOWN = [
        'oltp_read_write', 'oltp_write_only', 'oltp_read_only',
        'oltp_point_select', 'oltp_update_index', 'oltp_update_non_index',
        'oltp_random_points', 'oltp_random_ranges',
    ]
    name_lower = name.lower()
    for k in KNOWN:
        if k in name_lower:
            return k
    # DB 短格式兜底
    for short, full in SCEN_FROM_DB.items():
        if short in name:
            return full
    return None


def _extract_threads(test_name: str, results_key: str) -> Optional[int]:
    """从 test_name 或 results_key 中提取 threads。"""
    m = re.search(r'(\d+)\s*threads?', test_name or '', re.IGNORECASE)
    if m:
        return int(m.group(1))
    m = re.search(r'threads\s*=\s*(\d+)', results_key or '', re.IGNORECASE)
    if m:
        return int(m.group(1))
    return None


def _extract_arch(test_name: str) -> str:
    if '集中式' in (test_name or ''):
        return '集中式'
    if '分布式' in (test_name or ''):
        return '分布式'
    return ''


# ============== 适配器：本地 log ==============
def adapter_local_log(path: str) -> Dict[str, Any]:
    """复用 tdsql-b-whitepaper 的 parse_log。"""
    parse_log = _load_parse_log()
    header, recs = parse_log(path)
    product = header.get('engine', os.path.basename(path).split('.')[0])
    records = []
    for r in recs:
        records.append({
            'product': product,
            'scenario': f"oltp_{r['scenario'].replace('oltp_','').lstrip('_')}" if not r['scenario'].startswith('oltp_') else r['scenario'],
            'threads': r['threads'],
            'tps': r['tps'],
            'qps': r['qps'],
            'p95_ms': r['p95_ms'],
            'p99_ms': r.get('p99_ms'),
            'source_log': os.path.basename(path),
        })
    return {
        'meta': {
            'products': [product],
            'scenarios': sorted({r['scenario'] for r in records}),
            'concurrencies': sorted({r['threads'] for r in records}),
            'test_env': {},
            'log_meta': [{
                'product': product,
                'log_file': os.path.basename(path),
                'engine': header['engine'],
                'start_time': header['start_time'],
                'db_endpoint': header['db_endpoint'],
                'data_config': header['data_config'],
                'duration_sec': header['duration_sec'],
            }],
            'source_info': {'type': 'local_file', 'value': path, 'rows_fetched': len(records)},
        },
        'records': records,
    }


# ============== 适配器：本地 xlsx（沿用上游约定）==============
def adapter_local_xlsx(path: str) -> Dict[str, Any]:
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    product_sheets = [s for s in wb.sheetnames if not s.startswith('_')]
    expected = ['scenario', 'threads', 'tps', 'qps', 'p95_ms', 'p99_ms']
    records = []
    products = []
    for sn in product_sheets:
        ws = wb[sn]
        header = [c.value for c in ws[1]]
        if header[:6] != expected:
            raise RuntimeError(f'{sn} 表头不符合标准: {header}')
        products.append(sn)
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or row[0] is None:
                continue
            records.append({
                'product': sn,
                'scenario': row[0],
                'threads': row[1],
                'tps': row[2],
                'qps': row[3],
                'p95_ms': row[4],
                'p99_ms': row[5] if len(row) > 5 else None,
                'source_log': os.path.basename(path),
            })
    return {
        'meta': {
            'products': products,
            'scenarios': sorted({r['scenario'] for r in records}),
            'concurrencies': sorted({r['threads'] for r in records}),
            'test_env': {},
            'log_meta': [{'product': p, 'log_file': os.path.basename(path)} for p in products],
            'source_info': {'type': 'local_file', 'value': path, 'rows_fetched': len(records)},
        },
        'records': records,
    }


# ============== 适配器：本地 json ==============
def adapter_local_json(path: str) -> Dict[str, Any]:
    with open(path, encoding='utf-8') as f:
        data = json.load(f)
    if isinstance(data, dict) and 'records' in data and 'meta' in data:
        return data  # 已是标准格式
    if isinstance(data, list):
        return normalize_raw_rows(data, source_value=path, source_type='local_file')
    raise RuntimeError(f'JSON 格式不识别: {path}')


# ============== 适配器：本地 csv ==============
def adapter_local_csv(path: str) -> Dict[str, Any]:
    records = []
    with open(path, encoding='utf-8') as f:
        reader = csv.DictReader(f)
        for row in reader:
            records.append({
                'product': row.get('product', os.path.basename(path)),
                'scenario': row['scenario'],
                'threads': int(row['threads']),
                'tps': float(row['tps']),
                'qps': float(row['qps']),
                'p95_ms': float(row['p95_ms']) if row.get('p95_ms') else None,
                'p99_ms': float(row['p99_ms']) if row.get('p99_ms') else None,
            })
    return {
        'meta': {
            'products': sorted({r['product'] for r in records}),
            'scenarios': sorted({r['scenario'] for r in records}),
            'concurrencies': sorted({r['threads'] for r in records}),
            'test_env': {},
            'log_meta': [],
            'source_info': {'type': 'local_file', 'value': path, 'rows_fetched': len(records)},
        },
        'records': records,
    }


# ============== 数据种类检测 ==============
def _detect_data_kind(rows: List[Dict[str, Any]]) -> str:
    """根据本地原始测试行的字段识别数据种类。

    返回值：
      - 'tpcds_duration' : mrthree 工具 + tpcds_qNN/tpcds_load 场景（只有 duration_s）
      - 'sysbench_oltp'  : sysbench + oltp_* 场景（TPS/QPS/P95）
      - 'mixed'          : 混合或未知
    """
    if not rows:
        return 'unknown'
    tpcds_cnt = 0
    sysbench_cnt = 0
    for r in rows[:200]:
        dim = r.get('dimension') or {}
        if isinstance(dim, str):
            try:
                dim = json.loads(dim)
            except Exception:
                dim = {}
        test_name = (dim.get('test_name') or '').lower()
        test_scene = (dim.get('test_scene') or '').lower()
        tool = (dim.get('tool_name') or '').lower()
        if 'tpcds_q' in test_name or 'tpcds_q' in test_scene or 'tpcds_load' in test_name:
            tpcds_cnt += 1
        elif 'oltp_' in test_name or tool == 'sysbench':
            sysbench_cnt += 1
    if tpcds_cnt > sysbench_cnt and tpcds_cnt > 0:
        return 'tpcds_duration'
    if sysbench_cnt > 0:
        return 'sysbench_oltp'
    return 'mixed'


# ============== 适配器：本地导出 TPC-DS duration 行 ==============
_EXCLUDE_KEYWORDS = [
    'cleanup', 'clean_up', 'drop', 'truncate', 'setup', 'prepare', 'init',
    'create_table', 'load_data', 'create_database', 'import_data', 'insert_data',
    'download_test_case', 'precheck', 'sysbench_precheck', 'download',
]
_EXCLUDE_TOOLS = ['setup', 'mrthree_setup', 'precheck_tool']


def adapter_raw_rows_tpcds(rows: List[Dict[str, Any]], source_value: str = '', source_type: str = 'local_file') -> Dict[str, Any]:
    """TPC-DS 场景：每条 dimension.test_name 对应 1 个查询，results.duration_s 为耗时。"""
    import re as _re
    records = []
    env_info = {}
    for row in rows:
        dim = row.get('dimension') or {}
        results = row.get('results') or {}
        if isinstance(dim, str):
            try: dim = json.loads(dim)
            except Exception: dim = {}
        if isinstance(results, str):
            try: results = json.loads(results)
            except Exception: results = {}

        tool = (dim.get('tool_name') or '').lower()
        if tool in _EXCLUDE_TOOLS:
            continue
        test_name = dim.get('test_name') or ''
        tn_lower = test_name.lower()
        # 过滤数据准备/清理场景
        if any(k in tn_lower for k in _EXCLUDE_KEYWORDS):
            continue
        # 只保留 tpcds_q* 查询（load_Ng 仍保留，但单独打标签）
        if 'tpcds_q' not in tn_lower and 'tpcds_load' not in tn_lower:
            continue

        duration = _to_float(results.get('duration_s') or results.get('duration'))
        if duration is None:
            continue

        # 解析查询编号
        m = _re.search(r'tpcds_q(\d+)', tn_lower)
        q_no = int(m.group(1)) if m else None
        is_load = 'tpcds_load' in tn_lower

        deploy_arch = dim.get('deploy_arch') or ('standalone' if 'standalone' in tn_lower else
                                                 ('distributed' if 'distributed' in tn_lower else ''))
        product_type = dim.get('product_type') or dim.get('component_name') or 'unknown'
        product = f"{product_type}-{deploy_arch}" if deploy_arch else product_type

        records.append({
            'product': product,
            'query_no': q_no,
            'query_label': f'Q{q_no}' if q_no is not None else ('数据加载' if is_load else test_name),
            'test_name': test_name,
            'duration_s': duration,
            'is_load': is_load,
            'start_time': row.get('created'),
            'raw': {'dimension': dim, 'results_key': row.get('results_key')},
        })

        # 环境信息（取第一条）
        if not env_info:
            env_info = {
                'product_type': dim.get('product_type'),
                'deploy_arch': dim.get('deploy_arch'),
                'cn_version': dim.get('cn_version'),
                'dn_version': dim.get('dn_version'),
                'cvm_cpu': dim.get('cvm_cpu'),
                'cvm_memory': dim.get('cvm_memory'),
                'cpu_arch': dim.get('cpu_arch'),
                'machine_model': dim.get('machine_model'),
                'machine_type': dim.get('machine_type'),
                'cpu_performance_mode': dim.get('cpu_performance_mode'),
                'node_config': dim.get('node_config'),
                'network_type': dim.get('network_type'),
                'kernel_config': dim.get('kernel_config'),
                'requirement_type': dim.get('requirement_type'),
                'env_tag': dim.get('env_tag'),
                'test_category': dim.get('test_category'),
            }

    return {
        'meta': {
            'data_kind': 'tpcds_duration',
            'products': sorted({r['product'] for r in records}),
            'scenarios': sorted({r['query_label'] for r in records}),
            'concurrencies': [],
            'test_env': env_info,
            'log_meta': [{
                'product': (sorted({r['product'] for r in records}) or ['unknown'])[0],
                'engine': 'TPC-DS',
            }],
            'source_info': {
                'type': source_type,
                'value': source_value,
                'rows_fetched': len(records),
                'total_rows_in_source': len(rows),
            },
        },
        'records': records,
    }


# ============== 适配器：本地导出原始测试行转换（sysbench OLTP）==============
def adapter_raw_rows(rows: List[Dict[str, Any]], source_value: str = '', source_type: str = 'local_file') -> Dict[str, Any]:
    """本地导出的原始测试行 → 标准 records。"""
    records = []
    log_meta_set = {}

    for row in rows:
        dim = row.get('dimension') or {}
        results = row.get('results') or {}
        if isinstance(dim, str):
            dim = json.loads(dim)
        if isinstance(results, str):
            results = json.loads(results)

        test_name = dim.get('test_name', '')
        scenario = _normalize_scenario(test_name)
        threads = _extract_threads(test_name, row.get('results_key', ''))
        if not scenario or threads is None:
            continue  # 跳过无法解析的

        arch = _extract_arch(test_name)
        product_base = dim.get('tool_version') or dim.get('component_version') or 'unknown'
        product = f"{arch}-{product_base}" if arch else product_base

        # 提取关键指标（兼容导出行 #qps/#tps/#p95_latency_ms 及原始字段名）
        qps = _to_float(results.get('#qps') or results.get('qps') or results.get('QPS'))
        tps = _to_float(results.get('#tps') or results.get('tps') or results.get('TPS'))
        p95 = _to_float(results.get('#p95_latency_ms') or results.get('p95') or results.get('p95_latency') or results.get('latency_p95'))
        p99 = _to_float(results.get('#p99_latency_ms') or results.get('p99') or results.get('p99_latency') or results.get('latency_p99'))

        if qps is None or tps is None:
            # 主键点查/索引更新场景，TPS=QPS（每事务 1 语句）
            qps = qps or tps
            tps = tps or qps

        if qps is None or tps is None or p95 is None:
            continue

        records.append({
            'product': product,
            'scenario': scenario,
            'threads': threads,
            'tps': tps,
            'qps': qps,
            'p95_ms': p95,
            'p99_ms': p99,
            'start_time': row.get('created'),
            'raw': {'dimension': dim, 'results_key': row.get('results_key')},
        })

        # log_meta 去重
        key = product
        if key not in log_meta_set:
            log_meta_set[key] = {
                'product': product,
                'engine': dim.get('tool_name', ''),
                'start_time': row.get('created'),
                'data_config': row.get('results_key', ''),
                'duration_sec': _extract_duration(row.get('results_key', '')),
            }

    return {
        'meta': {
            'products': sorted({r['product'] for r in records}),
            'scenarios': sorted({r['scenario'] for r in records}),
            'concurrencies': sorted({r['threads'] for r in records}),
            'test_env': {},
            'log_meta': list(log_meta_set.values()),
            'source_info': {
                'type': source_type,
                'value': source_value,
                'rows_fetched': len(records),
            },
        },
        'records': records,
    }


def _to_float(v):
    if v is None:
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _extract_duration(rk: str) -> Optional[int]:
    m = re.search(r'time\s*=\s*(\d+)', rk or '')
    return int(m.group(1)) if m else None


# ============== 主分发 ==============
def load_records(intent: Dict[str, Any]) -> Dict[str, Any]:
    ds_type = intent['data_source_type']
    ds_value = intent.get('data_source_value')

    if ds_type == 'local_file':
        if not ds_value or not os.path.exists(ds_value):
            raise RuntimeError(f'E2001: 本地文件不存在: {ds_value}')
        ext = ds_value.rsplit('.', 1)[-1].lower()
        if ext == 'log':
            return adapter_local_log(ds_value)
        if ext == 'xlsx':
            return adapter_local_xlsx(ds_value)
        if ext == 'json':
            return adapter_local_json(ds_value)
        if ext == 'csv':
            return adapter_local_csv(ds_value)
        raise RuntimeError(f'E2002: 不识别的扩展名: {ext}')

    if ds_type == 'local_data':
        # 支持标准 records JSON、原始测试行单条 JSON，或多条数组
        data = json.loads(ds_value)
        if isinstance(data, dict) and 'records' in data and 'meta' in data:
            return data
        rows = data if isinstance(data, list) else [data]
        kind = _detect_data_kind(rows)
        if kind == 'tpcds_duration':
            return adapter_raw_rows_tpcds(rows, source_value='local_data', source_type='local_data')
        return adapter_raw_rows(rows, source_value='local_data', source_type='local_data')

    if ds_type == 'missing_data':
        raise RuntimeError('E2003: 缺少本地数据源，请提供 .log/.xlsx/.json/.csv 文件路径或粘贴 JSON 数据')

    raise RuntimeError(f'未知 data_source_type: {ds_type}')


def normalize_raw_rows(rows: List[Dict[str, Any]], source_value: str, source_type: str = 'local_file') -> Dict[str, Any]:
    """将本地导出的原始测试行转换为标准 records。"""
    if not rows:
        raise RuntimeError('E2004: 本地原始数据为空')
    kind = _detect_data_kind(rows)
    if kind == 'tpcds_duration':
        return adapter_raw_rows_tpcds(rows, source_value=source_value, source_type=source_type)
    return adapter_raw_rows(rows, source_value=source_value, source_type=source_type)


# ============== 数据质量门控 ==============
def quality_check(extracted: Dict[str, Any]) -> Dict[str, Any]:
    records = extracted['records']
    n = len(records)
    if n == 0:
        raise RuntimeError('E2004: records 为空')

    kind = extracted.get('meta', {}).get('data_kind', 'sysbench_oltp')

    if kind == 'tpcds_duration':
        null_dur = sum(1 for r in records if r.get('duration_s') is None)
        if null_dur:
            raise RuntimeError(f'E2005: duration_s 空值率超标 {null_dur}/{n}')
        return {
            'records': n,
            'duration_null_rate': 0,
            'queries': sorted({r.get('query_label', '-') for r in records}),
            'data_kind': kind,
        }

    # sysbench_oltp 默认逻辑
    null_tps = sum(1 for r in records if r.get('tps') is None)
    null_qps = sum(1 for r in records if r.get('qps') is None)
    null_p95 = sum(1 for r in records if r.get('p95_ms') is None)
    null_p99 = sum(1 for r in records if r.get('p99_ms') is None)

    if null_tps or null_qps or null_p95:
        raise RuntimeError(
            f'E2005: 空值率超标 — TPS={null_tps}/{n}, QPS={null_qps}/{n}, P95={null_p95}/{n}'
        )

    cbm = defaultdict(lambda: defaultdict(set))
    for r in records:
        cbm[r['product']][r['scenario']].add(r['threads'])

    return {
        'records': n,
        'tps_null_rate': null_tps / n,
        'qps_null_rate': null_qps / n,
        'p95_null_rate': null_p95 / n,
        'p99_null_rate': null_p99 / n,
        'concurrencies_by_product_scenario': {
            p: {s: sorted(ts) for s, ts in sd.items()} for p, sd in cbm.items()
        },
        'data_kind': kind,
    }


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--intent', required=True)
    ap.add_argument('--out', required=True)
    args = ap.parse_args()

    with open(args.intent, encoding='utf-8') as f:
        intent = json.load(f)

    extracted = load_records(intent)
    qc = quality_check(extracted)
    if qc.get('data_kind') != 'tpcds_duration':
        extracted['meta']['concurrencies_by_product_scenario'] = qc['concurrencies_by_product_scenario']

    os.makedirs(os.path.dirname(args.out) or '.', exist_ok=True)
    with open(args.out, 'w', encoding='utf-8') as f:
        json.dump(extracted, f, ensure_ascii=False, indent=2, default=str)
    print(f'✅ extracted_data.json 已保存: {args.out}（{qc["records"]} 条记录, kind={qc.get("data_kind","?")}）')
    print(f'   产品: {extracted["meta"]["products"]}')
    print(f'   场景: {extracted["meta"]["scenarios"][:10]}{"..." if len(extracted["meta"]["scenarios"])>10 else ""}')
    if qc.get('data_kind') != 'tpcds_duration':
        print(f'   并发档: {extracted["meta"]["concurrencies"]}')


if __name__ == '__main__':
    main()
