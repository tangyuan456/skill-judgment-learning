#!/usr/bin/env python3
"""
data_source_adapter.py — 统一数据接入（仅支持本地文件和粘贴数据）

适配器：
  - local_file: 本地 .log/.xlsx/.json/.csv
  - local_data: 粘贴的 JSON 文本
  - keyword_only: 仅关键词（将回退为提示用户提供文件）

标准结构：
{
  "meta": {"products": [...], "scenarios": [...], "concurrencies": [...], ...},
  "records": [{"product": ..., "scenario": ..., "threads": ..., "tps": ..., "qps": ..., "p95_ms": ...}, ...]
}
"""
import argparse
import csv
import json
import os
import re
import sys
from collections import defaultdict
from typing import Any, Dict, List, Optional

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from log_parser import parse_log  # noqa: E402


# ============== 反向场景映射 ==============
SCEN_FROM_DB = {
    'point_select': 'oltp_point_select',
    'read_only': 'oltp_read_only',
    'write_only': 'oltp_write_only',
    'read_write': 'oltp_read_write',
    'update_index': 'oltp_update_index',
    'update_non_index': 'oltp_update_non_index',
    'random_points': 'oltp_random_points',
    'random_ranges': 'oltp_random_ranges',
    'tpmC': 'tpmC',
}


def _normalize_scenario(name: str) -> Optional[str]:
    """从 test_name 字符串中识别 sysbench 标准场景名。"""
    if not name:
        return None
    m = re.search(r'(oltp_[a-z_]+)', name, re.IGNORECASE)
    if m:
        return m.group(1).lower()
    for short, full in SCEN_FROM_DB.items():
        if short in name:
            return full
    return None


def _extract_threads(test_name: str, results_key: str) -> Optional[int]:
    m = re.search(r'(\d+)\s*threads?', test_name or '', re.IGNORECASE)
    if m:
        return int(m.group(1))
    m = re.search(r'threads\s*=\s*(\d+)', results_key or '', re.IGNORECASE)
    if m:
        return int(m.group(1))
    return None


# ============== 适配器：本地 log ==============
def adapter_local_log(path: str) -> Dict[str, Any]:
    header, records_raw = parse_log(path)
    product = header.get('engine', os.path.basename(path).split('.')[0])
    records = []
    for r in records_raw:
        scenario = r['scenario']
        if not scenario.startswith('oltp_'):
            scenario = 'oltp_' + scenario.replace('oltp_', '').lstrip('_')
        records.append({
            'product': product,
            'scenario': scenario,
            'threads': r['threads'],
            'tps': r['tps'],
            'qps': r['qps'],
            'p95_ms': r.get('p95_ms'),
            'p99_ms': r.get('p99_ms'),
            'source_log': os.path.basename(path),
        })

    if not records:
        raise RuntimeError(f'日志文件未解析到有效数据: {path}')

    return {
        'meta': {
            'products': [product],
            'scenarios': sorted({r['scenario'] for r in records}),
            'concurrencies': sorted({r['threads'] for r in records}),
            'test_env': {},
            'log_meta': [{
                'product': product,
                'log_file': os.path.basename(path),
                'engine': header['engine'],
                'start_time': header['start_time'],
                'db_endpoint': header['db_endpoint'],
                'data_config': header['data_config'],
                'duration_sec': header['duration_sec'],
            }],
            'source_info': {'type': 'local_file', 'value': path, 'rows_fetched': len(records)},
        },
        'records': records,
    }


# ============== 适配器：本地 xlsx ==============
def adapter_local_xlsx(path: str) -> Dict[str, Any]:
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    product_sheets = [s for s in wb.sheetnames if not s.startswith('_')]
    expected = ['scenario', 'threads', 'tps', 'qps', 'p95_ms', 'p99_ms']
    records = []
    products = []
    for sn in product_sheets:
        ws = wb[sn]
        header = [c.value for c in ws[1]]
        if header[:6] != expected:
            raise RuntimeError(f'{sn} 表头不符合标准: {header}')
        products.append(sn)
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or row[0] is None:
                continue
            records.append({
                'product': sn,
                'scenario': row[0],
                'threads': row[1],
                'tps': row[2],
                'qps': row[3],
                'p95_ms': row[4],
                'p99_ms': row[5] if len(row) > 5 else None,
                'source_log': os.path.basename(path),
            })
    return {
        'meta': {
            'products': products,
            'scenarios': sorted({r['scenario'] for r in records}),
            'concurrencies': sorted({r['threads'] for r in records}),
            'test_env': {},
            'log_meta': [{'product': p, 'log_file': os.path.basename(path)} for p in products],
            'source_info': {'type': 'local_file', 'value': path, 'rows_fetched': len(records)},
        },
        'records': records,
    }


# ============== 适配器：本地 json ==============
def adapter_local_json(path: str) -> Dict[str, Any]:
    with open(path, encoding='utf-8') as f:
        data = json.load(f)
    if isinstance(data, dict) and 'records' in data and 'meta' in data:
        return data  # 已是标准格式
    if isinstance(data, list):
        # 列表格式：适配为 records
        return adapter_local_rows(data, source_value=path)
    raise RuntimeError(f'JSON 格式不识别: {path}')


# ============== 适配器：本地 csv ==============
def adapter_local_csv(path: str) -> Dict[str, Any]:
    records = []
    with open(path, encoding='utf-8-sig') as f:
        reader = csv.DictReader(f)
        for row in reader:
            records.append({
                'product': row.get('product', os.path.basename(path)),
                'scenario': row['scenario'],
                'threads': int(row['threads']),
                'tps': float(row['tps']),
                'qps': float(row['qps']),
                'p95_ms': float(row['p95_ms']) if row.get('p95_ms') else None,
                'p99_ms': float(row['p99_ms']) if row.get('p99_ms') else None,
            })
    return {
        'meta': {
            'products': sorted({r['product'] for r in records}),
            'scenarios': sorted({r['scenario'] for r in records}),
            'concurrencies': sorted({r['threads'] for r in records}),
            'test_env': {},
            'log_meta': [],
            'source_info': {'type': 'local_file', 'value': path, 'rows_fetched': len(records)},
        },
        'records': records,
    }


# ============== 适配器：本地 JSON 列表数据 ==============
def adapter_local_rows(rows: List[Dict[str, Any]], source_value: str = '') -> Dict[str, Any]:
    """将通用 JSON 行列表转换为标准 records 结构。"""
    records = []
    for row in rows:
        scenario = row.get('scenario', '') or row.get('test_name', '')
        if not scenario or not scenario.startswith('oltp_'):
            # 尝试从 test_name 中提取
            dim = row.get('dimension', {})
            if isinstance(dim, str):
                try:
                    dim = json.loads(dim)
                except Exception:
                    dim = {}
            test_name = dim.get('test_name', '') or row.get('test_name', '')
            scenario = _normalize_scenario(test_name) or scenario

        threads = row.get('threads') or _extract_threads(
            row.get('test_name', '') or (row.get('dimension', {}).get('test_name', '')
                                         if isinstance(row.get('dimension', {}), dict) else ''),
            row.get('results_key', ''),
        )

        tps = _to_float(row.get('tps') or row.get('TPS'))
        qps = _to_float(row.get('qps') or row.get('QPS'))
        results = row.get('results', {})
        if isinstance(results, str):
            try:
                results = json.loads(results)
            except Exception:
                results = {}
        p95 = _to_float(row.get('p95_ms') or results.get('p95') or results.get('p95_latency'))
        p99 = _to_float(row.get('p99_ms') or results.get('p99') or results.get('p99_latency'))

        if not scenario or threads is None:
            continue
        if tps is None and qps is None:
            continue
        qps = qps or tps
        tps = tps or qps

        product = row.get('product', source_value or 'test-product')
        records.append({
            'product': product,
            'scenario': scenario,
            'threads': threads,
            'tps': tps,
            'qps': qps,
            'p95_ms': p95,
            'p99_ms': p99,
        })

    if not records:
        raise RuntimeError(f'未能从数据中解析出有效记录: {source_value}')

    return {
        'meta': {
            'products': sorted({r['product'] for r in records}),
            'scenarios': sorted({r['scenario'] for r in records}),
            'concurrencies': sorted({r['threads'] for r in records}),
            'test_env': {},
            'log_meta': [],
            'source_info': {'type': 'local_data', 'value': source_value, 'rows_fetched': len(records)},
        },
        'records': records,
    }


def _to_float(v):
    if v is None:
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


# ============== 主分发 ==============
def load_records(intent: Dict[str, Any]) -> Dict[str, Any]:
    ds_type = intent['data_source_type']
    ds_value = intent.get('data_source_value')

    if ds_type == 'local_file':
        if not ds_value or not os.path.exists(ds_value):
            raise RuntimeError(f'本地文件不存在: {ds_value}')
        ext = ds_value.rsplit('.', 1)[-1].lower()
        if ext == 'log':
            return adapter_local_log(ds_value)
        if ext in ('xlsx', 'xls'):
            return adapter_local_xlsx(ds_value)
        if ext == 'json':
            return adapter_local_json(ds_value)
        if ext == 'csv':
            return adapter_local_csv(ds_value)
        raise RuntimeError(f'不识别的扩展名: {ext}')

    if ds_type == 'local_data':
        data = json.loads(ds_value)
        rows = data if isinstance(data, list) else [data]
        return adapter_local_rows(rows, source_value='local_data')

    if ds_type == 'keyword_only':
        raise RuntimeError(
            '未检测到具体的数据文件。'
            '请提供以下格式之一：\n'
            '  - sysbench .log 日志文件路径\n'
            '  - .xlsx Excel 文件路径\n'
            '  - .json 数据文件路径\n'
            '  - .csv 数据文件路径\n'
            '  - 直接粘贴 JSON 数据'
        )

    raise RuntimeError(f'未知 data_source_type: {ds_type}')


# ============== 数据质量门控 ==============
def quality_check(extracted: Dict[str, Any]) -> Dict[str, Any]:
    records = extracted['records']
    n = len(records)
    if n == 0:
        raise RuntimeError('records 为空')

    null_tps = sum(1 for r in records if r.get('tps') is None)
    null_qps = sum(1 for r in records if r.get('qps') is None)
    null_p95 = sum(1 for r in records if r.get('p95_ms') is None)
    null_p99 = sum(1 for r in records if r.get('p99_ms') is None)

    if null_tps or null_qps or null_p95:
        raise RuntimeError(
            f'数据空值率超标 — TPS={null_tps}/{n}, QPS={null_qps}/{n}, P95={null_p95}/{n}'
        )

    cbm = defaultdict(lambda: defaultdict(set))
    for r in records:
        cbm[r['product']][r['scenario']].add(r['threads'])

    return {
        'records': n,
        'tps_null_rate': null_tps / n,
        'qps_null_rate': null_qps / n,
        'p95_null_rate': null_p95 / n,
        'p99_null_rate': null_p99 / n,
        'concurrencies_by_product_scenario': {
            p: {s: sorted(ts) for s, ts in sd.items()} for p, sd in cbm.items()
        },
    }


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--intent', required=True)
    ap.add_argument('--out', required=True)
    args = ap.parse_args()

    with open(args.intent, encoding='utf-8') as f:
        intent = json.load(f)

    extracted = load_records(intent)
    qc = quality_check(extracted)
    extracted['meta']['concurrencies_by_product_scenario'] = qc['concurrencies_by_product_scenario']

    os.makedirs(os.path.dirname(args.out) or '.', exist_ok=True)
    with open(args.out, 'w', encoding='utf-8') as f:
        json.dump(extracted, f, ensure_ascii=False, indent=2, default=str)
    print(f'extracted_data.json 已保存: {args.out}（{qc["records"]} 条记录）')
    print(f'   产品: {extracted["meta"]["products"]}')
    print(f'   场景: {extracted["meta"]["scenarios"][:10]}{"..." if len(extracted["meta"]["scenarios"])>10 else ""}')
    print(f'   并发档: {extracted["meta"]["concurrencies"]}')


if __name__ == '__main__':
    main()
